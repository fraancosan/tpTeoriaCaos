from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, Alignment
import os
from datetime import datetime


class generadorExcel:
    def __init__(self) -> None:
      # Especifica el directorio donde se guardara el archivo Excel con las posibles soluciones
      self.directorio_resultados = "resultados"

      # Verifica si el directorio existe, de no existir se creara
      if not os.path.exists(self.directorio_resultados):
          os.makedirs(self.directorio_resultados)


      # Crear un nuevo libro de Excel
      self.libro_excel = Workbook()
      self.hoja_excel = self.libro_excel.active
      self.hoja_excel.title = f"Resultados {datetime.now().strftime('%d-%m-%Y %H-%M-%S')}"

      self.bordeDelgado = Side(border_style="thin", color="000000")

      # Instrucciones generales para todos los pasos
      # Encabezados
      self.hoja_excel['A6'] = f"Ejecucion Nro"
      self.hoja_excel['A6'].font = Font(bold=True)
      self.alinearCelda(self.hoja_excel['A6'])
      self.ponerBorde(self.hoja_excel['A6'], self.bordeDelgado)

      self.hoja_excel['B6'].font = Font(bold=True)
      self.hoja_excel['B6'] = f"Tasa de Crecimiento"
      self.alinearCelda(self.hoja_excel['B6'])
      self.ponerBorde(self.hoja_excel['B6'], self.bordeDelgado)

      self.hoja_excel['C6'] = f"Fitness"
      self.hoja_excel['C6'].font = Font(bold=True)
      self.alinearCelda(self.hoja_excel['C6'])
      self.ponerBorde(self.hoja_excel['C6'], self.bordeDelgado)

      # Tamaño columnas
      self.hoja_excel.column_dimensions["A"].width = 20
      self.hoja_excel.column_dimensions["B"].width = 25
      self.hoja_excel.column_dimensions["C"].width = 25

    def alinearCelda(self, celda):
        celda.alignment = Alignment(horizontal='center')

    def ponerBorde(self, celda, borde):
        celda.border = Border(top=borde, left=borde, right=borde, bottom=borde)

    def generar(self, datos, ciclos, caracteres, individuos, probabilidadMutacion, probabilidadCruce):
        # Informacion general
        self.hoja_excel['A1'] = f"Cantidad de ciclos: {ciclos}"
        self.hoja_excel['A2'] = f"Cantidad de genes por individuo: {caracteres}"
        self.hoja_excel['A3'] = f"Cantidad de individuos por generacion: {individuos}"
        self.hoja_excel['A4'] = f"Probabilidad de mutacion: {probabilidadMutacion}%"
        self.hoja_excel['A5'] = f"Probabilidad de cruce: {probabilidadCruce}%"

        # Se añade estilo a las celdas
        self.hoja_excel['A1'].font = Font(bold=True)
        self.alinearCelda(self.hoja_excel['A1'])
        self.ponerBorde(self.hoja_excel['A1'], self.bordeDelgado)

        self.hoja_excel['A2'].font = Font(bold=True)
        self.alinearCelda(self.hoja_excel['A2'])
        self.ponerBorde(self.hoja_excel['A2'], self.bordeDelgado)

        self.hoja_excel['A3'].font = Font(bold=True)
        self.alinearCelda(self.hoja_excel['A3'])
        self.ponerBorde(self.hoja_excel['A3'], self.bordeDelgado)

        self.hoja_excel['A4'].font = Font(bold=True)
        self.alinearCelda(self.hoja_excel['A4'])
        self.ponerBorde(self.hoja_excel['A4'], self.bordeDelgado)

        self.hoja_excel['A5'].font = Font(bold=True)
        self.alinearCelda(self.hoja_excel['A5'])
        self.ponerBorde(self.hoja_excel['A5'], self.bordeDelgado)

        # Se hace merge a las celdas
        self.hoja_excel.merge_cells('A1:C1')
        self.hoja_excel.merge_cells('A2:C2')
        self.hoja_excel.merge_cells('A3:C3')
        self.hoja_excel.merge_cells('A4:C4')
        self.hoja_excel.merge_cells('A5:C5')

        # Se llena el excel con los datos
        for ejecucion in range(len(datos)):
            self.hoja_excel[f"A{ejecucion+7}"] = f"Ejecucion {ejecucion+1}"
            self.hoja_excel[f"B{ejecucion+7}"] = datos[ejecucion].flotante
            self.hoja_excel[f"C{ejecucion+7}"] = datos[ejecucion].valorFitness

            # Se le da estilo a las celdas
            self.alinearCelda(self.hoja_excel[f"A{ejecucion+7}"])
            self.alinearCelda(self.hoja_excel[f"B{ejecucion+7}"])
            self.alinearCelda(self.hoja_excel[f"C{ejecucion+7}"])

            self.ponerBorde(self.hoja_excel[f"A{ejecucion+7}"], self.bordeDelgado)
            self.ponerBorde(self.hoja_excel[f"B{ejecucion+7}"], self.bordeDelgado)
            self.ponerBorde(self.hoja_excel[f"C{ejecucion+7}"], self.bordeDelgado)
        
        # Se ponen unas notas al final
        self.hoja_excel[f"A{len(datos)+7}"] = "Mientras mas cerca este el fitness del 1, mas cerca esta de la solucion"
        self.hoja_excel[f"A{len(datos)+8}"] = "Hay muchas soluciones posibles"

        # Se les da estilo
        self.hoja_excel[f"A{len(datos)+7}"].font = Font(bold=True)
        self.alinearCelda(self.hoja_excel[f"A{len(datos)+7}"])
        self.ponerBorde(self.hoja_excel[f"A{len(datos)+7}"], self.bordeDelgado)

        self.hoja_excel[f"A{len(datos)+8}"].font = Font(bold=True)
        self.alinearCelda(self.hoja_excel[f"A{len(datos)+8}"])
        self.ponerBorde(self.hoja_excel[f"A{len(datos)+8}"], self.bordeDelgado)

        # Se hace merge a las celdas
        self.hoja_excel.merge_cells(f"A{len(datos)+7}:C{len(datos)+7}")
        self.hoja_excel.merge_cells(f"A{len(datos)+8}:C{len(datos)+8}")

        # Se guarda el archivo
        self.libro_excel.save(f"{self.directorio_resultados}/resultados {datetime.now().strftime('%d-%m-%Y %H-%M-%S')}.xlsx")

        print(f"Se guardo el archivo en {self.directorio_resultados}/resultados {datetime.now().strftime('%d-%m-%Y %H-%M-%S')}.xlsx")